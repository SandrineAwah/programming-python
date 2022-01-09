'''
NAME:SANDRINE AWAH

python assignment.

openpyxl is a module in python that helps us interact with Excel(.xlxs ) files

the 11 and 12 line is to get access to my excel employeedata excel file'''

import csv
'''csv is a module in python that helps interact with csv files.
A csv file is a delimited text file that uses comma to separate values in which each line of 
the line of the file is data recorded'''
from openpyxl import workbook,load_workbook #iport workbork from openpyxl modulo
wb = load_workbook('employeedata.xlsx') #open workbook and give the file name

ws= wb.active # get a work book active sheet
range = ws["B2":"B31"]
'''this define my range at its column. it ranges from 2 to 31 and belong
to column B of the sheet'''

def employeedatabase():# defining a function
  '''from line 24 to 26 is simple print out the content in column B of the employeedata.xlsx sheet 
  (the email addressess)'''
  for cell in range:
    for x in cell:
        print(x.value)
employeedatabase()
print("******************************************************** Changed*****************************************")

def updated_employeedatabase():# defining a function
  for cell in range:
    for x in cell:
    
      text =x.value
           
      changesufix = text.replace("helpinghands.cm","handsinhands.org") 
      # the replace method is use to replace a text in another text
      x.value = changesufix 
    print(changesufix)#the new sufix is printed out
    wb.save('updateddatabase.xlsx')# it creates a new file with  new sufix(handsinhands.org)
updated_employeedatabase()

#***csv****
#line 46 help us to read the csv file "r" permit the file to be read
text = open("employeedata.csv","r")
  
# updating the column value/data(helpinghan.cmd)
text = ''.join([k for k in text])
# the join method takes all lines of a csv file an iterable and join them into one string
text = text.replace('helpinghands.cm','handsinhands.org')  
 # the replace method is use to replace a text in another text
x = open("updateddatabase","w")#the file is open in the write mode using "w"
# all the replace test is written in the updateddatabase.csv file
x.writelines(text)
x.close()


  




  
